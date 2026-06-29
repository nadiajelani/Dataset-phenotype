#!/usr/bin/env python3
"""
Locomotion dataset reconnaissance tool (lightweight, fast pass).

This is step 1 before running the full video_inventory.py probe. It does
NOT open or analyse video content, and does not transmit anything anywhere
-- it only reads local filenames, file sizes, dates, and lightweight
metadata from SLEAP (.slp/.h5) and spreadsheet (.csv/.xlsx) files, writing
a local report. Designed to run in minutes, not hours, even across a
large multi-year archive.

What it does:
  1. Maps the full folder tree (depth-limited) with file counts by extension
  2. Inspects every .slp / .h5 file: tries to read SLEAP metadata
     (video source, number of labelled frames, skeleton/keypoint names,
     number of tracks) if h5py is available; falls back to size/date only
     if h5py is not installed.
  3. Peeks at every .csv / .xlsx file: reads just the header row (and a
     couple of sample rows for .csv) without loading the full file, and
     flags any file whose headers contain likely structural-score or
     animal-ID terms.
  4. Writes a single text report plus a CSV summary table.

Requirements:
    pip install openpyxl
    pip install h5py        (optional -- script degrades gracefully without it)

Usage:
    python3 recon.py "C:\\Users\\njelani\\ownCloud\\Locomotion Project"
    (or the equivalent path on macOS/Linux)
"""

import sys
import os
import csv
import json
from pathlib import Path
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import load_workbook
    HAVE_OPENPYXL = True
except ImportError:
    HAVE_OPENPYXL = False

try:
    import h5py
    HAVE_H5PY = True
except ImportError:
    HAVE_H5PY = False

VIDEO_EXTS = {".mp4", ".avi", ".mov", ".mkv", ".m4v", ".wmv", ".mts", ".mxf"}

# Terms worth flagging in spreadsheet headers -- these are the columns that
# would actually answer the open RQ3/RQ4 question (does a structural score
# exist and is it linked to an animal/video?).
SCORE_TERMS = ["rh", "rear leg", "rear_leg", "rs", "rear side", "structural",
               "score", "bcs", "frame score", "muscle score", "claw", "foot angle",
               "fa", "fc", "msa", "carcase", "carcass"]
ID_TERMS = ["animal", "id", "tag", "nlis", "ear tag", "eartag", "rfid"]


def human_size(num_bytes):
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if abs(num_bytes) < 1024.0:
            return f"{num_bytes:.1f} {unit}"
        num_bytes /= 1024.0
    return f"{num_bytes:.1f} PB"


def scan_tree(root):
    """Walk the tree once, recording per-folder file counts/extensions/sizes."""
    root = Path(root)
    folder_stats = defaultdict(lambda: defaultdict(lambda: [0, 0]))  # folder -> ext -> [count, bytes]
    all_files = {"slp": [], "h5": [], "csv": [], "xlsx": [], "xls": []}

    file_count = 0
    for dirpath, dirnames, filenames in os.walk(root):
        rel_dir = os.path.relpath(dirpath, root)
        top_level = rel_dir.split(os.sep)[0] if rel_dir != "." else "(root)"
        for fname in filenames:
            file_count += 1
            fpath = Path(dirpath) / fname
            ext = fpath.suffix.lower().lstrip(".")
            try:
                size = fpath.stat().st_size
            except OSError:
                size = 0
            folder_stats[top_level][ext][0] += 1
            folder_stats[top_level][ext][1] += size

            if ext in all_files:
                all_files[ext].append(fpath)

            if file_count % 2000 == 0:
                print(f"  ...scanned {file_count} files so far")

    return folder_stats, all_files, file_count


def inspect_h5_or_slp(path):
    """Try to pull useful SLEAP metadata out of a .slp or .h5 file.
    SLEAP .slp files are HDF5 under the hood, so the same reader works for both.
    Falls back to size/date if h5py is unavailable or the file structure is unexpected."""
    info = {
        "file": str(path),
        "size": human_size(path.stat().st_size) if path.exists() else "?",
        "modified": datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d") if path.exists() else "?",
        "video_source": "", "n_frames_labelled": "", "n_tracks": "",
        "node_names": "", "note": ""
    }
    if not HAVE_H5PY:
        info["note"] = "h5py not installed -- only size/date available. Run inside your SLEAP env (~/.sleap_env) for full metadata."
        return info

    try:
        with h5py.File(path, "r") as f:
            # SLEAP stores video path(s) usually under 'video0/source_video' or similar attrs;
            # structure varies by SLEAP version, so we probe a few common locations defensively.
            top_keys = list(f.keys())
            info["note"] = f"top-level keys: {', '.join(top_keys[:8])}"

            if "frames" in f:
                try:
                    info["n_frames_labelled"] = f["frames"].shape[0]
                except Exception:
                    pass
            if "tracks_json" in f:
                info["n_tracks"] = "present (tracks_json found)"
            if "node_names" in f:
                try:
                    names = f["node_names"][:]
                    decoded = [n.decode() if isinstance(n, bytes) else str(n) for n in names]
                    info["node_names"] = ", ".join(decoded[:12])
                except Exception:
                    pass
            for vid_key in ["video0", "videos_json"]:
                if vid_key in f:
                    try:
                        attrs = dict(f[vid_key].attrs) if hasattr(f[vid_key], "attrs") else {}
                        if attrs:
                            info["video_source"] = str(attrs)[:200]
                    except Exception:
                        pass
    except Exception as e:
        info["note"] = f"could not open as HDF5: {e}"

    return info


def peek_csv(path, max_rows=3):
    try:
        with open(path, "r", encoding="utf-8", errors="replace") as fh:
            reader = csv.reader(fh)
            header = next(reader, [])
            samples = []
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                samples.append(row)
            return header, samples
    except Exception as e:
        return [], [f"ERROR: {e}"]


def peek_xlsx(path):
    if not HAVE_OPENPYXL:
        return [], "openpyxl not installed"
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        first_sheet = wb[sheet_names[0]]
        header = []
        for row in first_sheet.iter_rows(min_row=1, max_row=1, values_only=True):
            header = [str(c) if c is not None else "" for c in row]
        wb.close()
        return header, f"sheets: {', '.join(sheet_names)}"
    except Exception as e:
        return [], f"ERROR: {e}"


def flag_header(header):
    header_lower = " | ".join(str(h).lower() for h in header)
    score_hits = [t for t in SCORE_TERMS if t in header_lower]
    id_hits = [t for t in ID_TERMS if t in header_lower]
    return score_hits, id_hits


def main():
    if len(sys.argv) != 2:
        print('Usage: python3 recon.py "C:\\path\\to\\Locomotion Project"')
        sys.exit(1)

    root = sys.argv[1]
    if not os.path.isdir(root):
        print(f"ERROR: {root} is not a valid directory")
        sys.exit(1)

    print(f"Scanning folder tree under: {root}")
    print("(This reads filenames and sizes only -- no video files are opened)\n")
    folder_stats, all_files, total_files = scan_tree(root)
    print(f"\nDone scanning. {total_files} files found.\n")

    report_lines = []
    report_lines.append("=" * 70)
    report_lines.append("RECONNAISSANCE REPORT")
    report_lines.append(f"Root: {root}")
    report_lines.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    report_lines.append(f"h5py available: {HAVE_H5PY}  |  openpyxl available: {HAVE_OPENPYXL}")
    report_lines.append("=" * 70)

    report_lines.append("\n--- FOLDER / EXTENSION BREAKDOWN ---\n")
    for folder, exts in sorted(folder_stats.items(), key=lambda x: -sum(v[0] for v in x[1].values())):
        total = sum(v[0] for v in exts.values())
        total_bytes = sum(v[1] for v in exts.values())
        report_lines.append(f"{folder}  ({total} files, {human_size(total_bytes)})")
        for ext, (count, size) in sorted(exts.items(), key=lambda x: -x[1][0]):
            tag = " <- VIDEO" if f".{ext}" in VIDEO_EXTS else ""
            report_lines.append(f"    .{ext:<8} {count:>6}  {human_size(size):>10}{tag}")
        report_lines.append("")

    report_lines.append("\n--- SLEAP FILES (.slp / .h5) ---\n")
    slp_h5_files = all_files["slp"] + all_files["h5"]
    if not slp_h5_files:
        report_lines.append("None found.")
    else:
        report_lines.append(f"{len(slp_h5_files)} files found. Inspecting (this may take a moment)...")
        if not HAVE_H5PY:
            report_lines.append("NOTE: h5py is not installed in this Python environment.")
            report_lines.append("If you have a SLEAP environment set up (e.g. ~/.sleap_env), re-run")
            report_lines.append("this script using that environment's Python to get full metadata")
            report_lines.append("(labelled frame counts, keypoint/node names, video source paths).")
            report_lines.append("")
        for fpath in slp_h5_files[:50]:  # cap detailed inspection to first 50 to keep this fast
            info = inspect_h5_or_slp(fpath)
            report_lines.append(f"  {fpath.name}  [{info['size']}, modified {info['modified']}]")
            if info["node_names"]:
                report_lines.append(f"      keypoints/nodes: {info['node_names']}")
            if info["n_frames_labelled"]:
                report_lines.append(f"      labelled frames: {info['n_frames_labelled']}")
            if info["video_source"]:
                report_lines.append(f"      video source: {info['video_source']}")
            if info["note"]:
                report_lines.append(f"      note: {info['note']}")
        if len(slp_h5_files) > 50:
            report_lines.append(f"  ... and {len(slp_h5_files) - 50} more (capped detailed inspection at 50 for speed)")

    report_lines.append("\n--- SPREADSHEETS (.csv / .xlsx / .xls) ---\n")
    sheet_files = all_files["csv"] + all_files["xlsx"] + all_files["xls"]
    if not sheet_files:
        report_lines.append("None found.")
    else:
        flagged = []
        for fpath in sheet_files:
            ext = fpath.suffix.lower()
            if ext == ".csv":
                header, _ = peek_csv(fpath)
                extra = ""
            else:
                header, extra = peek_xlsx(fpath)
            score_hits, id_hits = flag_header(header)
            if score_hits or id_hits:
                flagged.append((fpath, header, score_hits, id_hits, extra))

        report_lines.append(f"{len(sheet_files)} spreadsheet files found.")
        report_lines.append(f"{len(flagged)} contain headers matching structural-score or animal-ID terms:\n")
        for fpath, header, score_hits, id_hits, extra in flagged:
            rel = os.path.relpath(fpath, root)
            report_lines.append(f"  {rel}")
            report_lines.append(f"      headers: {header[:15]}")
            if score_hits:
                report_lines.append(f"      ** SCORE-RELATED TERMS FOUND: {score_hits}")
            if id_hits:
                report_lines.append(f"      ** ID-RELATED TERMS FOUND: {id_hits}")
            if extra:
                report_lines.append(f"      ({extra})")
            report_lines.append("")

        if not flagged:
            report_lines.append("  No obvious matches -- doesn't mean scores aren't there, just that")
            report_lines.append("  column names didn't match the keyword list. Worth a manual look")
            report_lines.append("  at the .xlsx files listed above, especially in the 'ALL' folder.")

    report_lines.append("\n--- SUGGESTED NEXT STEPS ---\n")
    report_lines.append("1. Read through the folder/extension breakdown above. Folders named")
    report_lines.append("   'ALL' are often curated master sets -- check that one closely.")
    report_lines.append("2. Open the flagged spreadsheets directly and confirm whether they")
    report_lines.append("   link animal ID -> structural score -> video filename.")
    report_lines.append("3. If h5py wasn't available here, re-run this script from inside your")
    report_lines.append("   SLEAP environment to get keypoint names and labelled-frame counts")
    report_lines.append("   from the .slp/.h5 files -- this tells you if prior annotation work")
    report_lines.append("   is reusable for hock tracking specifically.")
    report_lines.append("4. Only after the above, run video_inventory.py -- and point it at")
    report_lines.append("   the specific subfolder(s) identified as relevant, not the full")
    report_lines.append("   812GB archive, to keep the run time reasonable.")

    report_text = "\n".join(str(line) for line in report_lines)
    out_path = "recon_report.txt"
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(report_text)

    print(report_text)
    print(f"\n\nFull report also saved to: {out_path}")


if __name__ == "__main__":
    main()