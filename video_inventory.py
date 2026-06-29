#!/usr/bin/env python3
"""
Locomotion video dataset inventory tool.

Run this LOCALLY against the folder where your OwnCloud locomotion videos
have been synced or downloaded. It does not upload, transmit, or share any
video content or filenames anywhere -- it only reads local file metadata
with ffprobe and writes a local spreadsheet. This keeps the data exactly
where Hassan's access conditions require it to stay.

Requirements:
    pip install openpyxl
    ffmpeg/ffprobe must be installed and on your PATH
        (macOS: brew install ffmpeg)

Usage:
    python3 video_inventory.py /path/to/locomotion/videos
b
Output:
    video_inventory.xlsx written to the current directory, with one row
    per video file found (recursively) under the path you give it.
"""

import sys
import os
import re
import json
import subprocess
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is not installed. Run: pip install openpyxl")
    sys.exit(1)

VIDEO_EXTENSIONS = {".mp4", ".mov", ".avi", ".mkv", ".m4v", ".wmv", ".mts", ".mxf"}

# Filename heuristics -- adjust these patterns if your actual filenames
# follow a different convention. The script flags anything it can't
# confidently parse rather than guessing wrong.
VIEW_KEYWORDS = {
    "rear": ["rear", "back", "behind", "posterior", "rh", "rs"],
    "side": ["side", "lateral", "profile"],
    "front": ["front", "fwd", "anterior", "head"],
}

ANGLE_KEYWORDS = {
    "parallel_perpendicular": ["perp", "parallel", "square", "straight"],
    "angled": ["angle", "angled", "oblique", "diag"],
}

ID_PATTERN = re.compile(r"(?:^|[_\-\s])([A-Z]{0,3}\d{3,6}[A-Z]?)(?=[_\-\s.]|$)")
YEAR_PATTERN = re.compile(r"^(19|20)\d{2}$")


def find_video_files(root):
    root = Path(root)
    files = []
    for path in root.rglob("*"):
        if path.is_file() and path.suffix.lower() in VIDEO_EXTENSIONS:
            files.append(path)
    return sorted(files)


def guess_view(name_lower):
    matches = [view for view, kws in VIEW_KEYWORDS.items() if any(kw in name_lower for kw in kws)]
    if len(matches) == 1:
        return matches[0]
    elif len(matches) > 1:
        return "ambiguous (" + ",".join(matches) + ")"
    return "unknown - check manually"


def guess_angle(name_lower):
    matches = [angle for angle, kws in ANGLE_KEYWORDS.items() if any(kw in name_lower for kw in kws)]
    if len(matches) == 1:
        return matches[0]
    elif len(matches) > 1:
        return "ambiguous"
    return "not in filename - check manually"


def guess_animal_id(stem):
    candidates = ID_PATTERN.findall(stem)
    # Filter out anything that looks like a 4-digit year (common false positive
    # since many filenames embed a recording date alongside the real animal ID)
    candidates = [c for c in candidates if not YEAR_PATTERN.match(c)]
    if not candidates:
        return "not detected - check manually"
    if len(candidates) > 1:
        return f"ambiguous ({', '.join(candidates)}) - check manually"
    return candidates[0]


def probe_video(path):
    """Run ffprobe and extract resolution, fps, duration, codec. Returns dict."""
    cmd = [
        "ffprobe", "-v", "error",
        "-select_streams", "v:0",
        "-show_entries", "stream=width,height,r_frame_rate,codec_name,nb_frames",
        "-show_entries", "format=duration,size",
        "-of", "json",
        str(path)
    ]
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
        if result.returncode != 0:
            return {"error": f"ffprobe failed: {result.stderr.strip()[:200]}"}
        data = json.loads(result.stdout)
        stream = data.get("streams", [{}])[0] if data.get("streams") else {}
        fmt = data.get("format", {})

        width = stream.get("width", "")
        height = stream.get("height", "")
        codec = stream.get("codec_name", "")

        fps_raw = stream.get("r_frame_rate", "")
        fps = ""
        if fps_raw and "/" in fps_raw:
            try:
                num, den = fps_raw.split("/")
                fps = round(float(num) / float(den), 2) if float(den) != 0 else ""
            except (ValueError, ZeroDivisionError):
                fps = fps_raw

        duration_raw = fmt.get("duration", "")
        duration = round(float(duration_raw), 2) if duration_raw else ""

        size_bytes = fmt.get("size", "")
        size_mb = round(int(size_bytes) / (1024 * 1024), 2) if size_bytes else ""

        return {
            "width": width, "height": height, "fps": fps,
            "duration_sec": duration, "codec": codec, "size_mb": size_mb,
            "error": ""
        }
    except subprocess.TimeoutExpired:
        return {"error": "ffprobe timed out"}
    except Exception as e:
        return {"error": f"unexpected error: {e}"}


def get_file_mtime(path):
    try:
        return datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d")
    except Exception:
        return ""


def build_inventory(root_path):
    files = find_video_files(root_path)
    if not files:
        print(f"No video files found under {root_path}")
        print(f"Looked for extensions: {', '.join(sorted(VIDEO_EXTENSIONS))}")
        return []

    print(f"Found {len(files)} video files. Probing metadata (this may take a while for large folders)...")
    rows = []
    for i, path in enumerate(files, 1):
        if i % 25 == 0 or i == len(files):
            print(f"  ...{i}/{len(files)}")

        name_lower = path.name.lower()
        meta = probe_video(path)

        rows.append({
            "filename": path.name,
            "relative_folder": str(path.parent.relative_to(root_path)) if path.parent != Path(root_path) else "(root)",
            "guessed_view": guess_view(name_lower),
            "guessed_camera_angle": guess_angle(name_lower),
            "guessed_animal_id": guess_animal_id(path.stem),
            "width": meta.get("width", ""),
            "height": meta.get("height", ""),
            "fps": meta.get("fps", ""),
            "duration_sec": meta.get("duration_sec", ""),
            "codec": meta.get("codec", ""),
            "size_mb": meta.get("size_mb", ""),
            "file_modified_date": get_file_mtime(path),
            "probe_error": meta.get("error", ""),
            "annotation_linked": "",   # fill in manually once you check Zhi/Muhammad's annotation set
            "structural_score_linked": "",  # fill in manually -- the binding RH-availability question
            "notes": "",
        })
    return rows


def write_workbook(rows, output_path):
    FONT = "Arial"
    HEADER_FILL = PatternFill("solid", start_color="1F3864")
    HEADER_FONT = Font(name=FONT, bold=True, color="FFFFFF", size=10)
    FLAG_FILL = PatternFill("solid", start_color="FFF2CC")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    fields = ["filename", "relative_folder", "guessed_view", "guessed_camera_angle",
              "guessed_animal_id", "width", "height", "fps", "duration_sec", "codec",
              "size_mb", "file_modified_date", "probe_error", "annotation_linked",
              "structural_score_linked", "notes"]
    widths = [34, 22, 22, 22, 18, 8, 8, 7, 11, 9, 9, 14, 24, 16, 20, 30]

    wb = Workbook()
    ws = wb.active
    ws.title = "Video inventory"

    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for col_idx, field in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field.replace("_", " ").title())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for r_idx, row in enumerate(rows, start=2):
        for c_idx, field in enumerate(fields, start=1):
            value = row.get(field, "")
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            needs_review = (
                (field == "guessed_view" and "unknown" in str(value)) or
                (field == "guessed_camera_angle" and "check manually" in str(value)) or
                (field == "guessed_animal_id" and "check manually" in str(value)) or
                (field == "probe_error" and value)
            )
            if needs_review:
                cell.fill = FLAG_FILL

    # Summary sheet
    summary = wb.create_sheet("Summary")
    summary.column_dimensions["A"].width = 32
    summary.column_dimensions["B"].width = 14
    total = len(rows)
    by_view = {}
    by_angle = {}
    errors = 0
    for row in rows:
        by_view[row["guessed_view"]] = by_view.get(row["guessed_view"], 0) + 1
        by_angle[row["guessed_camera_angle"]] = by_angle.get(row["guessed_camera_angle"], 0) + 1
        if row["probe_error"]:
            errors += 1

    summary_rows = [("Total video files found", total), ("Files with ffprobe errors", errors), ("", "")]
    summary_rows.append(("By guessed view (from filename):", ""))
    for k, v in sorted(by_view.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {k}", v))
    summary_rows.append(("", ""))
    summary_rows.append(("By guessed camera angle (from filename):", ""))
    for k, v in sorted(by_angle.items(), key=lambda x: -x[1]):
        summary_rows.append((f"  {k}", v))
    summary_rows.append(("", ""))
    summary_rows.append(("IMPORTANT:", ""))
    summary_rows.append(("Filename-based guesses are NOT reliable.", ""))
    summary_rows.append(("Rows highlighted yellow in 'Video inventory' need", ""))
    summary_rows.append(("manual review -- open a sample of those clips directly", ""))
    summary_rows.append(("to confirm view angle and camera alignment by eye.", ""))

    for r_idx, (label, value) in enumerate(summary_rows, start=1):
        c1 = summary.cell(row=r_idx, column=1, value=label)
        c1.font = Font(name=FONT, size=10, bold=label.endswith(":"))
        summary.cell(row=r_idx, column=2, value=value).font = Font(name=FONT, size=10)

    wb.save(output_path)


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 video_inventory.py ")
        sys.exit(1)

    root_path = sys.argv[1]
    if not os.path.isdir(root_path):
        print(f"ERROR: {root_path} is not a valid directory")
        sys.exit(1)

    rows = build_inventory(root_path)
    if not rows:
        sys.exit(0)

    output_path = "video_inventory.xlsx"
    write_workbook(rows, output_path)
    print(f"\nDone. Wrote {len(rows)} rows to {output_path}")
    print("\nNext steps:")
    print("  1. Open video_inventory.xlsx -- rows highlighted yellow need manual review")
    print("  2. Spot-check a sample of 'rear' + 'parallel_perpendicular' guesses by eye")
    print("     to confirm the ~195 figure Hassan mentioned")
    print("  3. Fill in 'Annotation Linked' and 'Structural Score Linked' columns")
    print("     once you've checked Zhi/Muhammad's annotation set against these filenames")


if __name__ == "__main__":
    main()